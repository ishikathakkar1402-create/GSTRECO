import streamlit as st
import pandas as pd
from rapidfuzz import fuzz
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from fpdf import FPDF

# --- Branding ---
BRAND_NAME = "RECONIX"
LOGO_PATH = r"C:\Users\barodapc01\Desktop\Adv IT 28_i\LOGO.jpeg"
QR_PATH = r"C:\Users\barodapc01\Desktop\Adv IT 28_i\qrCODE.jpeg"

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# --- Styling (Roboto everywhere + dark blue background) ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Roboto&display=swap');
html, body, [class*="css"], .stMarkdown, .stTextInput, .stButton, .stDownloadButton, .stDataFrame {
    font-family: 'Roboto', sans-serif !important;
}
.stApp {
    background: linear-gradient(135deg, #1e3c72, #2a5298); /* Dark blue gradient */
}
h1, h2, h3, h4 {
    text-decoration: underline;
}
</style>
""", unsafe_allow_html=True)

# --- Intelligent Matching ---
def intelligent_match(row, df2):
    for _, r in df2.iterrows():
        score_gstin = fuzz.ratio(str(row["Supplier GSTIN"]), str(r["Supplier GSTIN"]))
        score_name = fuzz.ratio(str(row["Supplier Name"]), str(r["Supplier Name"]))
        score_invoice = fuzz.ratio(str(row["Invoice Number"]), str(r["Invoice Number"]))
        score_date = fuzz.ratio(str(row["Invoice Date"]), str(r["Invoice Date"]))
        if score_gstin == 100 and score_invoice == 100 and score_date == 100:
            return "Matched"
        elif score_gstin > 90:
            return "GSTIN fuzzy match"
        elif score_name > 85:
            return "Name fuzzy match"
        elif score_invoice > 85:
            return "Invoice fuzzy match"
        elif score_date > 85:
            return "Date variation"
    return "Not matched"

# --- Login Page ---
def login_page():
    st.image(LOGO_PATH, width=120)
    st.title(f"Welcome to {BRAND_NAME}")
    user = st.text_input("User ID")
    pwd = st.text_input("Password", type="password")
    if st.button("Login"):
        if user == "ISHIKA" and pwd == "ISHIKA":
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Invalid credentials")

# --- Dashboard ---
def dashboard_page():
    st.sidebar.image(LOGO_PATH, width=100)
    st.sidebar.title(f"{BRAND_NAME} Dashboard")

    page = st.radio("Navigate", [
        "⚙️ Run Reconciliation", "📈 Drill‑Down", "👥 Assign Users", "⬇️ Download Reports"
    ])

    purchase_file = st.file_uploader("Upload Purchase Register (Excel)", type=["xlsx"])
    gstr2b_file = st.file_uploader("Upload GSTR‑2B (Excel)", type=["xlsx"])

    if purchase_file and gstr2b_file:
        purchase_df = pd.read_excel(purchase_file)
        gstr2b_df = pd.read_excel(gstr2b_file)

        purchase_df.columns = purchase_df.columns.str.strip()
        gstr2b_df.columns = gstr2b_df.columns.str.strip()

        # Merge PR and GSTR-2B
        same_entries = pd.merge(
            purchase_df, gstr2b_df,
            on=["Supplier GSTIN","Supplier Name","Invoice Number","Invoice Type","Invoice Date"],
            suffixes=("_PR","_2B")
        )

        # Duplicate key fields to create PR and 2B versions
        same_entries["Supplier GSTIN_PR"] = same_entries["Supplier GSTIN"]
        same_entries["Supplier GSTIN_2B"] = same_entries["Supplier GSTIN"]
        same_entries["Supplier Name_PR"] = same_entries["Supplier Name"]
        same_entries["Supplier Name_2B"] = same_entries["Supplier Name"]
        same_entries["Invoice Number_PR"] = same_entries["Invoice Number"]
        same_entries["Invoice Number_2B"] = same_entries["Invoice Number"]
        same_entries["Invoice Type_PR"] = same_entries["Invoice Type"]
        same_entries["Invoice Type_2B"] = same_entries["Invoice Type"]
        same_entries["Invoice Date_PR"] = same_entries["Invoice Date"]
        same_entries["Invoice Date_2B"] = same_entries["Invoice Date"]

        # Add difference columns
        same_entries["Taxable Diff"] = same_entries["Taxable Value_PR"] - same_entries["Taxable Value_2B"]
        same_entries["IGST Diff"] = same_entries["IGST_PR"] - same_entries["IGST_2B"]
        same_entries["CGST Diff"] = same_entries["CGST_PR"] - same_entries["CGST_2B"]
        same_entries["SGST Diff"] = same_entries["SGST_PR"] - same_entries["SGST_2B"]

        # Add remarks
        same_entries["Remarks"] = same_entries.apply(lambda x: intelligent_match(x, gstr2b_df), axis=1)

        # Reorder columns exactly as specified
        same_entries = same_entries[
            [
                "Supplier GSTIN_PR","Supplier GSTIN_2B",
                "Supplier Name_PR","Supplier Name_2B",
                "Invoice Number_PR","Invoice Number_2B",
                "Invoice Type_PR","Invoice Type_2B",
                "Invoice Date_PR","Invoice Date_2B",
                "Taxable Value_PR","IGST_PR","CGST_PR","SGST_PR",
                "Taxable Value_2B","IGST_2B","CGST_2B","SGST_2B",
                "Taxable Diff","IGST Diff","CGST Diff","SGST Diff",
                "Remarks"
            ]
        ]

        only_purchase = purchase_df.merge(gstr2b_df, on=["Supplier GSTIN","Supplier Name","Invoice Number","Invoice Type","Invoice Date"], how="left", indicator=True)
        only_purchase = only_purchase[only_purchase["_merge"]=="left_only"]

        only_2b = gstr2b_df.merge(purchase_df, on=["Supplier GSTIN","Supplier Name","Invoice Number","Invoice Type","Invoice Date"], how="left", indicator=True)
        only_2b = only_2b[only_2b["_merge"]=="left_only"]

        analysis = pd.DataFrame({
            "Purchase Register Total": [purchase_df["Taxable Value"].sum()],
            "GSTR-2B Total": [gstr2b_df["Taxable Value"].sum()],
            "Difference": [purchase_df["Taxable Value"].sum() - gstr2b_df["Taxable Value"].sum()],
            "Matched Count": [len(same_entries)],
            "Only in Purchase Count": [len(only_purchase)],
            "Only in 2B Count": [len(only_2b)]
        })

        # --- Run Reconciliation ---
        if page == "⚙️ Run Reconciliation":
            st.subheader("Run Reconciliation")
            if st.button("Run"):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    same_entries.to_excel(writer, sheet_name="Same Entries", index=False)
                    only_purchase.to_excel(writer, sheet_name="Only in Purchase", index=False)
                    only_2b.to_excel(writer, sheet_name="Only in 2B", index=False)
                    analysis.to_excel(writer, sheet_name="Data Analysis", index=False)
                output.seek(0)
                wb = load_workbook(output)

                # Formatting
                header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark Blue
                header_font = Font(name="Rockwell Nova", size=10, bold=True, color="FFFFFF")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                rockwell_font = Font(name="Rockwell Nova", size=10)

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    ws.insert_rows(1)
                    ws["A1"] = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    for cell in ws[2]:
                        cell.fill = header_fill
                        cell.font = header_font
                    for row in ws.iter_rows(min_row=3):
                        for cell in row:
                            cell.font = rockwell_font
                            cell.border = thin_border
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max_length + 2

                # Visuals sheet
                ws_chart = wb.create_sheet("Visuals")
                ws_chart.append(["Category", "Value"])
                ws_chart.append(["Purchase Register", purchase_df["Taxable Value"].sum()])
                ws_chart.append(["GSTR-2B", gstr2b_df["Taxable Value"].sum()])

                bar = BarChart()
                data = Reference(ws_chart, min_col=2, min_row=2, max_row=3)
                cats = Reference(ws_chart, min_col=1, min_row=2, max_row=3)
                bar.add_data(data, titles_from_data=False)
                bar.set_categories(cats)
                bar.title = "Taxable Value Comparison"
                ws_chart.add_chart(bar, "E5")

                line = LineChart()
                line.add_data(data, titles_from_data=False)
                line.set_categories(cats)
                line.title = "Trend Line"
                ws_chart.add_chart(line, "E20")

                # Save Excel
                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                st.session_state["excel_report"] = final_output

                # PDF Report (replica of Excel with formatting)
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.image(LOGO_PATH, 10, 8, 33)
                pdf.cell(200, 10, txt=f"{BRAND_NAME} Report", ln=True, align="C")
                pdf.cell(200, 10, txt=f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")

                # Render each Excel sheet into PDF
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    pdf.ln(10)
                    pdf.set_font("Arial", size=12)
                    pdf.cell(200, 10, txt=f"Sheet: {sheet}", ln=True)
                    pdf.set_font("Arial", size=8)

                    # Header row
                    header = [str(cell.value) if cell.value else "" for cell in ws[2]]
                    pdf.set_fill_color(0, 0, 139)  # Dark Blue
                    pdf.set_text_color(255, 255, 255)
                    pdf.cell(200, 8, txt=" | ".join(header), ln=True, fill=True)

                    # Reset text color
                    pdf.set_text_color(0, 0, 0)

                    # Data rows
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        pdf.cell(200, 6, txt=row_text, ln=True)

                # Watermark
                pdf.set_text_color(200,200,200)
                pdf.set_font("Arial", size=50)
                pdf.text(60,150,BRAND_NAME)

                pdf_output = pdf.output(dest="S").encode("latin-1")
                st.session_state["pdf_report"] = pdf_output

                st.success("Reconciliation complete! Go to 'Download Reports' tab to get your files.")

        # --- Drill‑Down ---
        if page == "📈 Drill‑Down":
            st.subheader("Pivot & Charts")
            pivot = purchase_df.pivot_table(index="Supplier Name", values="Taxable Value", aggfunc="sum")
            st.dataframe(pivot)
            st.bar_chart(pivot)

        # --- Assign Users ---
        if page == "👥 Assign Users":
            st.subheader("Add Users & Payment QR")
            new_user = st.text_input("Enter User Name")
            new_id = st.text_input("Enter User ID")
            new_pass = st.text_input("Enter Password", type="password")
            if st.button("Add User"):
                st.success(f"User {new_user} with ID {new_id} added!")
            st.image(QR_PATH, caption="Scan to Pay", width=200)

        # --- Download Reports ---
        if page == "⬇️ Download Reports":
            st.subheader("Download Reports")

            # Path selector
            save_path = st.text_input("Enter folder path to save reports", value=r"C:\Users\barodapc01\Desktop")

            if "excel_report" in st.session_state:
                file_name_xlsx = f"{save_path}\\gst_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    "📊 Download Excel Report",
                    data=st.session_state["excel_report"],
                    file_name=file_name_xlsx,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            if "pdf_report" in st.session_state:
                file_name_pdf = f"{save_path}\\gst_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                st.download_button(
                    "📄 Download PDF Report",
                    data=st.session_state["pdf_report"],
                    file_name=file_name_pdf,
                    mime="application/pdf"
                )

    # Footer branding
    st.markdown(
        f"<div style='position:fixed; bottom:10px; right:10px; color:white;'>Made by Ishika</div>",
        unsafe_allow_html=True
    )

# --- App Flow ---
if not st.session_state.authenticated:
    login_page()
else:
    dashboard_page()
